#! python3
# getOpenWeather.py - Prints the weather for a location from the command line

import openpyxl

import json
import requests
import sys

from textMyself import textmyself

APPID = 'b5960608a0956435089e99ba4a34699a'

def get_weather_data(location):
    # Download the JSON data from OpenWeatherMap.org's API
    url = f'https://api.openweathermap.org/data/2.5/forecast?q={location}&appid={APPID}'
    response = requests.get(url)

    # Uncomment to see the raw JSON text:
    # print(response.text)

    # Load JSON data into a Python variable
    weatherData = json.loads(response.text)

    # Print the entire weatherData for debugging
    print("Complete weather data:")
    print(json.dumps(weatherData, indent=2))

    return weatherData

def print_weather_description(weatherData):
    try:
        w = weatherData['list']
        print(f'Current weather in {location}:')
        print(w[0]['weather'][0]['main'], '-', w[0]['weather'][0]['description'])
        print()
        print('Tomorrow:')
        print(w[1]['weather'][0]['main'], '-', w[1]['weather'][0]['description'])
        print()
        print('Day after tomorrow:')
        print(w[2]['weather'][0]['main'], '-', w[2]['weather'][0]['description'])
    except KeyError as e:
        print(f"Error: {e}. Please check the structure of the 'weatherData' variable.")

def save_to_excel(weatherData, excel_file):
    try:
        w = weatherData['list']

        # Create a new Excel workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Add headers
        sheet['A1'] = 'Date'
        sheet['B1'] = 'Main Weather'
        sheet['C1'] = 'Description'

        # Add data
        for i in range(3):  # Consider adding more data if needed
            sheet.cell(row=i + 2, column=1, value=w[i]['dt_txt'])
            sheet.cell(row=i + 2, column=2, value=w[i]['weather'][0]['main'])
            sheet.cell(row=i + 2, column=3, value=w[i]['weather'][0]['description'])

        # Save the workbook to the specified file
        workbook.save(excel_file)

        print(f'Data saved to {excel_file}')
    except KeyError as e:
        print(f"Error: {e}. Please check the structure of the 'weatherData' variable.")

if __name__ == '__main__':
    # Compute location from command line arguments
    if len(sys.argv) < 2:
        print('Usage: getOpenWeather.py city_name, 2-letter_country_code')
        sys.exit()

    location = ' '.join(sys.argv[1:])
    weather_data = get_weather_data(location)
    print_weather_description(weather_data)

    # Send a SMS of the weather data
    textmyself(f'current weather in {location} has been updated')

    # Save weather data to Excel
    save_to_excel(weather_data, 'weather_data.xlsx')
